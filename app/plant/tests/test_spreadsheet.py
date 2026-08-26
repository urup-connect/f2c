"""Reading a filled-in template, with no database anywhere.

``spreadsheet`` issues no query, which is what lets every coercion and every
refusal below be exercised against a workbook built in memory. These are the
tests that matter most in the whole upload path, because a spreadsheet is
untrusted input from a person under time pressure and the failure modes are all
quiet:

**A date read wrong by a month.** ``03/04/2026`` is April in Johannesburg and
March in Chicago. Guessing produces a harvest estimate nobody questions until a
member asks where their plant is, so the reader refuses a string it cannot read
unambiguously.

**A price rounded rather than refused.** ``100.005`` becoming ``100.01`` puts a
cent on a member's bill that nobody authorised.

**A duplicate inside the file.** The unique constraint would catch it on the
second insert, partway through a batch, with a message naming an index.

**A truncated report.** A cultivator whose date column is formatted as text has
one mistake on every row. The reader must find all of them, and the command must
say when it stopped printing.
"""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook

from ..spreadsheet import (
    COLUMNS,
    HEADINGS,
    MAX_ROWS,
    PLANTS_SHEET,
    REFERENCE_SHEET,
    SheetError,
    build_template,
    read_rows,
)

GOOD_ROW = {
    'cultivator_plant_id': 'POT-1',
    'strain': 'OG Kush',
    'grow_price': 950,
    'planting_date': date(2026, 3, 1),
    'estimated_bloom_date': date(2026, 4, 30),
    'estimated_harvest_date': date(2026, 6, 29),
    'minimum_yield_grams': 30,
    'finished_product_types': None,
    'batch': None,
}


def workbook_bytes(rows, *, headings=None, sheet_title=PLANTS_SHEET):
    """A workbook in memory, from dicts keyed like :data:`COLUMNS`."""
    book = Workbook()
    sheet = book.active
    sheet.title = sheet_title

    keys = [key for key, _, _ in COLUMNS]
    sheet.append(headings if headings is not None else [HEADINGS[k] for k in keys])
    for row in rows:
        sheet.append([row.get(key) for key in keys])

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream


def row(**overrides):
    return GOOD_ROW | overrides


class TemplateTests(SimpleTestCase):
    def test_the_template_has_a_plants_sheet_and_a_reference_sheet(self):
        book = load_workbook(_saved(build_template([('OG Kush', 'Pre-rolls')])))

        self.assertIn(PLANTS_SHEET, book.sheetnames)
        self.assertIn(REFERENCE_SHEET, book.sheetnames)

    def test_the_headings_are_the_ones_the_reader_looks_for(self):
        """Otherwise the template the platform hands out is one it refuses."""
        book = load_workbook(_saved(build_template()))
        sheet = book[PLANTS_SHEET]

        headings = [cell.value for cell in sheet[1]]
        self.assertEqual(headings, [heading for _, heading, _ in COLUMNS])

    def test_a_round_trip_of_the_template_reads_back(self):
        """The strongest test here: what `build_template` writes is what
        `read_rows` accepts, so the two cannot drift apart."""
        book = build_template([('OG Kush', 'Pre-rolls')])
        sheet = book[PLANTS_SHEET]
        sheet.append([
            'POT-1', 'OG Kush', 950,
            date(2026, 3, 1), date(2026, 4, 30), date(2026, 6, 29),
            30, None, None,
        ])

        rows, errors = read_rows(_saved(book))

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['cultivator_plant_id'], 'POT-1')

    def test_the_reference_sheet_lists_the_cultivators_strains(self):
        book = load_workbook(_saved(build_template([
            ('OG Kush', 'Pre-rolls'), ('Durban Poison', 'Pre-rolls, Loose'),
        ])))
        sheet = book[REFERENCE_SHEET]

        self.assertEqual(sheet['A2'].value, 'OG Kush')
        self.assertEqual(sheet['B3'].value, 'Pre-rolls, Loose')

    def test_the_strain_column_gets_a_dropdown(self):
        """The single most common upload error is a strain name that is not one.
        A dropdown stops it being made rather than reporting it afterwards."""
        book = build_template([('OG Kush', 'Pre-rolls')])

        validations = book[PLANTS_SHEET].data_validations.dataValidation
        self.assertEqual(len(validations), 1)
        self.assertIn(REFERENCE_SHEET, validations[0].formula1)

    def test_no_dropdown_when_the_cultivator_has_no_listings(self):
        """An empty list would be a dropdown with nothing in it, which Excel
        renders as a cell that refuses every value."""
        book = build_template([])

        self.assertEqual(book[PLANTS_SHEET].data_validations.dataValidation, [])

    def test_the_template_explains_itself(self):
        """The file is what gets forwarded to whoever does the typing."""
        book = load_workbook(_saved(build_template()))
        notes = ' '.join(
            str(cell.value or '') for cell in book[REFERENCE_SHEET]['D']
        )

        self.assertIn('2026-03-01', notes)
        self.assertIn('two decimal places', notes)

    def test_there_is_no_column_for_anything_the_platform_generates(self):
        """`cultivator-stock-upload.md` separates the fields a cultivator supplies
        from the ones the system calculates. A serial column would invite a
        cultivator to fill it in."""
        headings = ' '.join(heading for _, heading, _ in COLUMNS).casefold()

        for generated in ('serial', 'pseudonym', 'leaf rating', 'status', 'days'):
            with self.subTest(generated=generated):
                self.assertNotIn(generated, headings)

    def test_there_is_no_cultivator_column(self):
        """The brief lists "Cultivator ID" as a field. It is deliberately not a
        column: it would let one cultivator load stock as another."""
        keys = {key for key, _, _ in COLUMNS}

        self.assertNotIn('cultivator', keys)
        self.assertNotIn('cultivator_id', keys)


class SheetShapeTests(SimpleTestCase):
    def test_a_file_that_is_not_a_workbook_is_refused_in_one_sentence(self):
        with self.assertRaises(SheetError) as refused:
            read_rows(BytesIO(b'not a spreadsheet'))

        self.assertIn('.xlsx', str(refused.exception))

    def test_a_workbook_with_no_plants_sheet_is_refused(self):
        with self.assertRaises(SheetError) as refused:
            read_rows(workbook_bytes([row()], sheet_title='Sheet1'))

        self.assertIn(PLANTS_SHEET, str(refused.exception))

    def test_a_missing_required_column_is_not_reported_as_a_bad_row(self):
        """It is the wrong file. Reporting "row 2: strain is required" nine
        hundred times hides the real cause."""
        headings = [
            HEADINGS[key] for key, _, _ in COLUMNS if key != 'strain'
        ]
        with self.assertRaises(SheetError) as refused:
            read_rows(workbook_bytes([], headings=headings))

        self.assertIn('Strain', str(refused.exception))

    def test_columns_may_be_reordered(self):
        """Found by heading, not by position, so a cultivator who moves a column
        has not broken the file."""
        keys = [key for key, _, _ in COLUMNS]
        book = Workbook()
        sheet = book.active
        sheet.title = PLANTS_SHEET
        reversed_keys = list(reversed(keys))
        sheet.append([HEADINGS[key] for key in reversed_keys])
        sheet.append([GOOD_ROW.get(key) for key in reversed_keys])

        rows, errors = read_rows(_saved(book))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['strain'], 'OG Kush')

    def test_a_headings_only_workbook_is_refused(self):
        with self.assertRaises(SheetError) as refused:
            read_rows(workbook_bytes([]))

        self.assertIn('no plants', str(refused.exception))

    def test_blank_rows_are_skipped_not_reported(self):
        """Excel hands back trailing empties for any sheet somebody scrolled."""
        rows, errors = read_rows(workbook_bytes([
            row(), {}, {}, row(cultivator_plant_id='POT-2'),
        ]))

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 2)

    def test_too_many_rows_is_refused_rather_than_truncated(self):
        """A silent cap would load part of a file and report success."""
        with self.assertRaises(SheetError) as refused:
            read_rows(workbook_bytes([
                row(cultivator_plant_id=f'POT-{n}') for n in range(MAX_ROWS + 2)
            ]))

        self.assertIn('nothing was loaded', str(refused.exception))


class RequiredFieldTests(SimpleTestCase):
    def test_a_missing_required_value_names_the_column(self):
        rows, errors = read_rows(workbook_bytes([row(strain=None)]))

        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0].row, 2)
        self.assertEqual(errors[0].column, 'Strain')
        self.assertIn('required', errors[0].message)

    def test_the_row_number_is_the_one_excel_shows(self):
        """A report a cultivator has to translate is one they will misread."""
        rows, errors = read_rows(workbook_bytes([
            row(), row(cultivator_plant_id='POT-2', grow_price=None),
        ]))

        self.assertEqual([error.row for error in errors], [3])

    def test_every_missing_field_in_a_row_is_reported_at_once(self):
        rows, errors = read_rows(workbook_bytes([
            row(strain=None, grow_price=None, minimum_yield_grams=None)
        ]))

        self.assertEqual(
            {error.column for error in errors},
            {'Strain', 'Grow price (R)', 'Minimum yield (g)'},
        )

    def test_the_optional_columns_may_be_blank(self):
        rows, errors = read_rows(workbook_bytes([
            row(finished_product_types=None, batch=None)
        ]))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['finished_product_types'], [])
        self.assertEqual(rows[0]['batch'], '')


class DateTests(SimpleTestCase):
    def test_a_real_date_cell_is_read(self):
        rows, errors = read_rows(workbook_bytes([row()]))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['planting_date'], date(2026, 3, 1))

    def test_a_datetime_is_read_as_its_date(self):
        """Excel hands back datetimes for date-formatted cells."""
        rows, errors = read_rows(workbook_bytes([
            row(planting_date=datetime(2026, 3, 1, 14, 30))
        ]))

        self.assertEqual(rows[0]['planting_date'], date(2026, 3, 1))

    def test_an_iso_string_is_read(self):
        rows, errors = read_rows(workbook_bytes([row(planting_date='2026-03-01')]))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['planting_date'], date(2026, 3, 1))

    def test_an_ambiguous_date_is_refused_rather_than_guessed(self):
        """The failure this whole rule exists for. `03/04/2026` is April here and
        March in Chicago, and a planting date wrong by a month is a harvest
        estimate wrong by a month that nobody questions."""
        rows, errors = read_rows(workbook_bytes([row(planting_date='03/04/2026')]))

        self.assertEqual(rows, [])
        self.assertEqual(errors[0].column, 'Planting date')
        self.assertIn('2026-03-01', errors[0].message)

    def test_the_refusal_says_how_to_fix_it(self):
        _, errors = read_rows(workbook_bytes([row(planting_date='1 March 2026')]))

        self.assertIn('Format the cell as a date', errors[0].message)

    def test_bloom_before_planting_is_refused(self):
        rows, errors = read_rows(workbook_bytes([
            row(estimated_bloom_date=date(2026, 2, 1))
        ]))

        self.assertEqual(rows, [])
        self.assertIn('flower before it was planted', errors[0].message)

    def test_harvest_before_planting_is_refused(self):
        rows, errors = read_rows(workbook_bytes([
            row(
                estimated_bloom_date=date(2026, 1, 1),
                estimated_harvest_date=date(2026, 2, 1),
            )
        ]))

        self.assertEqual(rows, [])
        self.assertTrue(
            any('harvested before it was planted' in e.message for e in errors)
        )

    def test_transposed_bloom_and_harvest_dates_are_caught(self):
        """No query needed, so it is caught in the file rather than by a check
        constraint naming an index."""
        rows, errors = read_rows(workbook_bytes([
            row(
                estimated_bloom_date=date(2026, 6, 29),
                estimated_harvest_date=date(2026, 4, 30),
            )
        ]))

        self.assertEqual(rows, [])
        self.assertIn('before bloom', errors[0].message)


class NumberTests(SimpleTestCase):
    def test_an_integer_is_read_as_a_decimal(self):
        rows, _ = read_rows(workbook_bytes([row(grow_price=950)]))

        self.assertEqual(rows[0]['grow_price'], Decimal('950.00'))

    def test_a_two_decimal_price_is_read_exactly(self):
        rows, errors = read_rows(workbook_bytes([row(grow_price=1234.56)]))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['grow_price'], Decimal('1234.56'))

    def test_more_than_two_decimals_is_refused_rather_than_rounded(self):
        """Quietly making 100.005 into 100.01 puts a cent on a member's bill
        that nobody authorised."""
        rows, errors = read_rows(workbook_bytes([row(grow_price=100.005)]))

        self.assertEqual(rows, [])
        self.assertIn('two decimal places', errors[0].message)

    def test_a_thousands_separator_is_tolerated(self):
        """A cultivator typing 1,250 into a text-formatted cell has not made a
        mistake worth refusing."""
        rows, errors = read_rows(workbook_bytes([row(grow_price='1,250')]))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['grow_price'], Decimal('1250.00'))

    def test_words_are_refused(self):
        rows, errors = read_rows(workbook_bytes([row(grow_price='nine fifty')]))

        self.assertEqual(rows, [])
        self.assertIn('must be a number', errors[0].message)

    def test_zero_is_refused(self):
        rows, errors = read_rows(workbook_bytes([row(grow_price=0)]))

        self.assertEqual(rows, [])
        self.assertIn('more than zero', errors[0].message)

    def test_a_negative_price_is_refused(self):
        rows, errors = read_rows(workbook_bytes([row(grow_price=-950)]))

        self.assertEqual(rows, [])
        self.assertIn('more than zero', errors[0].message)

    def test_a_boolean_is_refused_rather_than_becoming_one_rand(self):
        """`True` is an int in Python, so without the guard a TRUE cell would
        load a plant priced at R1."""
        rows, errors = read_rows(workbook_bytes([row(grow_price=True)]))

        self.assertEqual(rows, [])
        self.assertIn('must be a number', errors[0].message)


class ProductTypeColumnTests(SimpleTestCase):
    def test_a_comma_separated_list_is_split(self):
        rows, errors = read_rows(workbook_bytes([
            row(finished_product_types='Pre-rolls, Loose')
        ]))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['finished_product_types'], ['Pre-rolls', 'Loose'])

    def test_stray_separators_are_ignored(self):
        rows, _ = read_rows(workbook_bytes([
            row(finished_product_types=' Pre-rolls , , Loose ,')
        ]))

        self.assertEqual(rows[0]['finished_product_types'], ['Pre-rolls', 'Loose'])


class DuplicateTests(SimpleTestCase):
    def test_the_same_plant_id_twice_in_one_file_is_caught(self):
        """The unique constraint would catch it on the second insert, partway
        through a batch, with a message naming an index.

        The *first* occurrence stands and the duplicate is the error, which is
        what lets the message point back at the row it collides with. Nothing is
        written either way -- the service refuses an upload with any error at all
        -- so which of the two is blamed only affects how readable the report is.
        """
        rows, errors = read_rows(workbook_bytes([row(), row()]))

        self.assertEqual([r['_row'] for r in rows], [2])
        self.assertEqual(errors[0].row, 3)
        self.assertIn('row 2', errors[0].message)

    def test_the_comparison_ignores_case(self):
        rows, errors = read_rows(workbook_bytes([
            row(cultivator_plant_id='POT-1'), row(cultivator_plant_id='pot-1'),
        ]))

        self.assertEqual([r['_row'] for r in rows], [2])
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0].row, 3)

    def test_one_plant_id_may_repeat_across_two_strains(self):
        """A cultivator numbering pots per greenhouse rather than per farm has
        not made a mistake."""
        rows, errors = read_rows(workbook_bytes([
            row(strain='OG Kush'), row(strain='Durban Poison'),
        ]))

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 2)


class ReportingTests(SimpleTestCase):
    def test_every_bad_row_is_reported_not_just_the_first(self):
        """A cultivator whose date column is formatted as text has the same
        mistake on every row and needs to know it is every row."""
        rows, errors = read_rows(workbook_bytes([
            row(cultivator_plant_id=f'POT-{n}', planting_date='03/04/2026')
            for n in range(5)
        ]))

        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 5)

    def test_good_rows_survive_beside_bad_ones(self):
        """So a dry run reports the count of what would load."""
        rows, errors = read_rows(workbook_bytes([
            row(cultivator_plant_id='POT-1'),
            row(cultivator_plant_id='POT-2', grow_price='free'),
            row(cultivator_plant_id='POT-3'),
        ]))

        self.assertEqual([r['cultivator_plant_id'] for r in rows], ['POT-1', 'POT-3'])
        self.assertEqual(len(errors), 1)

    def test_an_error_reads_as_a_sentence_about_a_cell(self):
        _, errors = read_rows(workbook_bytes([row(grow_price='free')]))

        rendered = str(errors[0])
        self.assertIn('Row 2', rendered)
        self.assertIn('Grow price', rendered)
        self.assertIn('free', rendered)


def _saved(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
