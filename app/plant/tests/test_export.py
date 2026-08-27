"""Exporting stock on hand.

``drawio``, cultivator story v1 asks for "SOH imports and exports" and, higher up
the same story, for two separate inventory screens — "My inventory for sale" and
"My member-owned inventory" — plus "late items". Those three things are what the
scopes and the overdue flag are, so they are what is asserted here.

Four of these tests are about something other than the columns:

**The default scope is unsold.** Stock on hand means what a cultivator still
holds, and an export that quietly included sold plants would overstate the farm
to whoever read it.

**One cultivator's export contains one cultivator's plants.** The same property
the upload has, on the way out.

**The owner column is a nickname, and is absent when nothing is owned.** An
export is a file that leaves the platform and gets forwarded. `member-roles.md`
conceals members behind a nickname and C19 is the open question about what a
cultivator may see of one at all, so the file carries the least that is useful.

**The day counts share one "today".** A report whose first row was computed a
second before midnight and its last a second after has two answers in it.
"""
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.core.management import CommandError, call_command
from django.test import TestCase
from openpyxl import load_workbook

from ..models import Batch, OwnershipReason, PlantStatus
from ..services import (
    SCOPE_ALL,
    SCOPE_FOR_SALE,
    SCOPE_MEMBER_OWNED,
    build_stock_export,
    stock_for_export,
    stock_rows,
    stock_summary,
)
from ..spreadsheet import EXPORT_COLUMNS, STOCK_SHEET, SUMMARY_SHEET
from .support import HARVESTS, PLANTED, PlantTestCase


def sheet_rows(workbook, title=STOCK_SHEET):
    sheet = workbook[title]
    headings = [cell.value for cell in sheet[1]]
    return headings, [
        dict(zip(headings, [cell.value for cell in row]))
        for row in sheet.iter_rows(min_row=2)
    ]


def saved(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return load_workbook(stream)


class ScopeTests(PlantTestCase):
    def setUp(self):
        super().setUp()
        self.unsold = self.make_plant(cultivator_plant_id='POT-1')
        self.sold = self.make_plant(cultivator_plant_id='POT-2')
        self.sold.transfer_to(self.member, reason=OwnershipReason.PURCHASE)
        self.withdrawn = self.make_plant(cultivator_plant_id='POT-3')
        self.withdrawn.disable()

    def test_stock_on_hand_is_unsold_by_default(self):
        """The default is the whole point: "stock on hand" is what the cultivator
        still holds, and including sold plants would overstate the farm."""
        plants = stock_for_export(self.cultivator)

        self.assertEqual([plant.pk for plant in plants], [self.unsold.pk])

    def test_member_owned_is_the_other_screen(self):
        plants = stock_for_export(self.cultivator, SCOPE_MEMBER_OWNED)

        self.assertEqual([plant.pk for plant in plants], [self.sold.pk])

    def test_all_is_everything_still_growing(self):
        plants = stock_for_export(self.cultivator, SCOPE_ALL)

        self.assertEqual(
            {plant.pk for plant in plants}, {self.unsold.pk, self.sold.pk}
        )

    def test_a_withdrawn_plant_is_in_no_scope(self):
        """It is not stock, and a report that counts it overstates the farm."""
        for scope in (SCOPE_FOR_SALE, SCOPE_MEMBER_OWNED, SCOPE_ALL):
            with self.subTest(scope=scope):
                plants = stock_for_export(self.cultivator, scope)
                self.assertNotIn(
                    self.withdrawn.pk, {plant.pk for plant in plants}
                )

    def test_another_cultivators_stock_is_not_in_the_export(self):
        """The upload's security property, on the way out."""
        other = self.another_member('other@example.com', 'Tygerberg')

        self.assertEqual(list(stock_for_export(other, SCOPE_ALL)), [])

    def test_an_unknown_scope_is_refused(self):
        with self.assertRaises(ValueError) as refused:
            stock_for_export(self.cultivator, 'everything')

        self.assertIn('not a scope', str(refused.exception))

    def test_the_order_is_the_one_a_grower_plans_by(self):
        """Harvest date first. A crop is worked in date order, not serial order."""
        later = self.make_plant(
            cultivator_plant_id='POT-4',
            estimated_harvest_date=HARVESTS + timedelta(days=30),
        )
        earlier = self.make_plant(
            cultivator_plant_id='POT-5',
            estimated_harvest_date=HARVESTS - timedelta(days=30),
        )

        plants = list(stock_for_export(self.cultivator))

        self.assertEqual(plants[0].pk, earlier.pk)
        self.assertEqual(plants[-1].pk, later.pk)


class RowTests(PlantTestCase):
    def test_a_row_carries_what_a_cultivator_cannot_work_out_themselves(self):
        """The mirror of the template: it omits the generated fields because a
        cultivator must not supply them, and this leads with them."""
        plant = self.make_plant(grow_price=Decimal('1650.00'))

        row = stock_rows([plant], today=PLANTED)[0]

        self.assertEqual(row['serial'], plant.serial)
        self.assertEqual(row['leaf_rating'], Decimal('1.5'))
        self.assertEqual(row['status'], 'Preflowering')
        self.assertEqual(row['days_to_harvest'], 120)
        self.assertEqual(row['strain'], 'OG Kush')
        self.assertEqual(row['finished_product_types'], 'Pre-rolls')

    def test_the_batch_is_its_reference_and_blank_when_there_is_none(self):
        batch = Batch.objects.create(
            cultivator=self.cultivator, reference='2026-01'
        )
        batched = self.make_plant(cultivator_plant_id='POT-1', batch=batch)
        loose = self.make_plant(cultivator_plant_id='POT-2')

        rows = stock_rows([batched, loose], today=PLANTED)

        self.assertEqual(rows[0]['batch'], '2026-01')
        self.assertEqual(rows[1]['batch'], '')

    def test_a_plant_past_its_estimated_harvest_is_flagged_overdue(self):
        """"Late items", which the cultivator story asks for by name. Invisible
        in a list sorted by date unless it is labelled."""
        plant = self.make_plant()

        row = stock_rows([plant], today=HARVESTS + timedelta(days=1))[0]

        self.assertEqual(row['overdue'], 'Yes')
        self.assertEqual(row['days_to_harvest'], -1)

    def test_a_plant_on_schedule_is_not_flagged(self):
        plant = self.make_plant()

        row = stock_rows([plant], today=PLANTED)[0]

        self.assertEqual(row['overdue'], '')

    def test_a_harvested_plant_is_never_overdue(self):
        """`days_to_harvest` stops counting once harvested, so the flag has to
        follow it rather than compare dates itself."""
        plant = self.make_plant()
        plant.mark_harvested(HARVESTS)

        row = stock_rows([plant], today=HARVESTS + timedelta(days=60))[0]

        self.assertEqual(row['overdue'], '')
        self.assertEqual(row['harvested_on'], HARVESTS)

    def test_every_row_shares_one_today(self):
        """Otherwise a file straddling midnight has two answers in it."""
        plants = [
            self.make_plant(cultivator_plant_id=f'POT-{n}') for n in range(3)
        ]

        rows = stock_rows(plants, today=PLANTED)

        self.assertEqual({row['days_to_harvest'] for row in rows}, {120})

    def test_the_owner_is_a_nickname_and_never_an_email_address(self):
        plant = self.make_plant()
        plant.transfer_to(self.member, reason=OwnershipReason.PURCHASE)

        row = stock_rows([plant], today=PLANTED)[0]

        self.assertEqual(row['held_by'], 'Sam')
        self.assertNotIn('@', row['held_by'])

    def test_unsold_stock_has_no_owner_to_show(self):
        row = stock_rows([self.make_plant()], today=PLANTED)[0]

        self.assertEqual(row['held_by'], '')


class SummaryTests(PlantTestCase):
    def test_it_totals_what_a_farm_asks_about(self):
        rows = stock_rows([
            self.make_plant(cultivator_plant_id='POT-1',
                            grow_price=Decimal('950.00')),
            self.make_plant(cultivator_plant_id='POT-2',
                            grow_price=Decimal('1650.00')),
        ], today=PLANTED)

        summary = stock_summary(rows)

        self.assertEqual(summary['plants'], 2)
        self.assertEqual(summary['grow_price_total'], Decimal('2600.00'))
        self.assertEqual(summary['yield_total'], Decimal('60.00'))
        self.assertEqual(summary['overdue'], 0)

    def test_it_counts_the_overdue(self):
        rows = stock_rows(
            [self.make_plant()], today=HARVESTS + timedelta(days=1)
        )

        self.assertEqual(stock_summary(rows)['overdue'], 1)

    def test_it_groups_by_status_and_by_strain(self):
        first = self.make_plant(cultivator_plant_id='POT-1')
        second = self.make_plant(cultivator_plant_id='POT-2')
        second.status = PlantStatus.IN_BLOOM
        second.save(update_fields=['status'])

        summary = stock_summary(stock_rows([first, second], today=PLANTED))

        self.assertEqual(
            summary['by_status'], [('In bloom', 1), ('Preflowering', 1)]
        )
        self.assertEqual(summary['by_strain'], [('OG Kush', 2)])

    def test_an_empty_export_totals_zero_rather_than_failing(self):
        summary = stock_summary([])

        self.assertEqual(summary['plants'], 0)
        self.assertEqual(summary['grow_price_total'], Decimal('0.00'))


class WorkbookTests(PlantTestCase):
    def test_the_workbook_has_a_stock_sheet_and_a_summary(self):
        book = saved(build_stock_export([self.make_plant()]))

        self.assertIn(STOCK_SHEET, book.sheetnames)
        self.assertIn(SUMMARY_SHEET, book.sheetnames)

    def test_a_row_per_plant(self):
        for n in range(3):
            self.make_plant(cultivator_plant_id=f'POT-{n}')

        book = saved(build_stock_export(stock_for_export(self.cultivator)))
        _, rows = sheet_rows(book)

        self.assertEqual(len(rows), 3)

    def test_the_owner_column_is_left_out_when_nothing_is_owned(self):
        """An empty column reads as missing data. Stock on hand is unsold by
        definition, so the usual export has nothing to put there."""
        self.make_plant()

        headings, _ = sheet_rows(saved(
            build_stock_export(stock_for_export(self.cultivator))
        ))

        self.assertNotIn('Held by', headings)

    def test_the_owner_column_appears_when_something_is_owned(self):
        plant = self.make_plant()
        plant.transfer_to(self.member, reason=OwnershipReason.PURCHASE)

        headings, rows = sheet_rows(saved(build_stock_export(
            stock_for_export(self.cultivator, SCOPE_MEMBER_OWNED)
        )))

        self.assertIn('Held by', headings)
        self.assertEqual(rows[0]['Held by'], 'Sam')

    def test_the_owner_column_can_be_forced_off(self):
        plant = self.make_plant()
        plant.transfer_to(self.member, reason=OwnershipReason.PURCHASE)

        headings, _ = sheet_rows(saved(build_stock_export(
            stock_for_export(self.cultivator, SCOPE_ALL), held_by=False
        )))

        self.assertNotIn('Held by', headings)

    def test_the_headings_are_the_export_columns(self):
        self.make_plant()

        headings, _ = sheet_rows(saved(
            build_stock_export(stock_for_export(self.cultivator))
        ))

        self.assertEqual(
            headings,
            [heading for key, heading, _ in EXPORT_COLUMNS if key != 'held_by'],
        )

    def test_the_summary_names_the_scope(self):
        """So a file found on a desktop later says what it was a report of."""
        self.make_plant()

        book = saved(build_stock_export(
            stock_for_export(self.cultivator), scope_label='Unsold — testing'
        ))
        values = [
            str(cell.value or '') for row in book[SUMMARY_SHEET] for cell in row
        ]

        self.assertIn('Unsold — testing', values)

    def test_an_empty_export_is_headings_and_a_summary(self):
        """A cultivator whose whole crop has sold gets a file, not an error."""
        book = saved(build_stock_export(stock_for_export(self.cultivator)))
        headings, rows = sheet_rows(book)

        self.assertTrue(headings)
        self.assertEqual(rows, [])

    def test_the_export_is_not_a_re_import(self):
        """Every plant in it already exists, so the duplicate check refuses it.
        The import half of "SOH imports and exports" is the template, which is
        for stock that is new."""
        from ..spreadsheet import PLANTS_SHEET, SheetError, read_rows

        self.make_plant()
        book = build_stock_export(stock_for_export(self.cultivator))

        stream = BytesIO()
        book.save(stream)
        stream.seek(0)

        with self.assertRaises(SheetError) as refused:
            read_rows(stream)

        # It does not even have the sheet an upload reads, which is the first and
        # bluntest way the two files differ.
        self.assertIn(PLANTS_SHEET, str(refused.exception))


class CommandTests(PlantTestCase):
    def export(self, directory, *args):
        from pathlib import Path

        destination = Path(directory) / 'stock.xlsx'
        call_command(
            'export_stock', 'grower@example.com',
            '--output', str(destination), *args
        )
        return destination

    def test_it_writes_a_workbook(self):
        import tempfile

        self.make_plant()

        with tempfile.TemporaryDirectory() as directory:
            destination = self.export(directory)

            self.assertTrue(destination.is_file())
            _, rows = sheet_rows(load_workbook(destination))
            self.assertEqual(len(rows), 1)

    def test_the_scope_is_honoured(self):
        import tempfile

        plant = self.make_plant()
        plant.transfer_to(self.member, reason=OwnershipReason.PURCHASE)

        with tempfile.TemporaryDirectory() as directory:
            for_sale = self.export(directory, '--scope', SCOPE_FOR_SALE)
            _, rows = sheet_rows(load_workbook(for_sale))
            self.assertEqual(rows, [])

            owned = self.export(directory, '--scope', SCOPE_MEMBER_OWNED)
            _, rows = sheet_rows(load_workbook(owned))
            self.assertEqual(len(rows), 1)

    def test_an_empty_export_still_writes_a_file(self):
        import tempfile

        with tempfile.TemporaryDirectory() as directory:
            destination = self.export(directory)

            self.assertTrue(destination.is_file())

    def test_an_account_that_is_not_a_cultivator_is_refused(self):
        with self.assertRaises(CommandError) as refused:
            call_command('export_stock', 'member@example.com')

        self.assertIn('not a cultivator', str(refused.exception))

    def test_an_unknown_scope_is_refused_by_the_parser(self):
        from django.core.management.base import CommandError as CE

        with self.assertRaises((CE, SystemExit)):
            call_command('export_stock', 'grower@example.com',
                         '--scope', 'everything')


class AdminActionTests(PlantTestCase):
    def test_it_returns_a_spreadsheet_of_the_selection(self):
        """The case the command cannot cover: whatever staff have filtered."""
        from django.contrib.admin.sites import AdminSite

        from ..admin import PlantAdmin
        from ..models import Plant

        self.make_plant(cultivator_plant_id='POT-1')
        self.make_plant(cultivator_plant_id='POT-2')

        admin = PlantAdmin(Plant, AdminSite())
        response = admin.export_stock(
            None, Plant.objects.filter(cultivator_plant_id='POT-1')
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('attachment', response['Content-Disposition'])

        book = load_workbook(BytesIO(response.content))
        _, rows = sheet_rows(book)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['Cultivator plant ID'], 'POT-1')
