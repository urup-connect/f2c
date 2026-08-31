"""The cultivator's stock template, and reading a filled-in copy of it back.

``twp-tasks/cultivator-stock-upload.md``: "Cultivators can load individual
plants or batch upload multiple plants using an excel template." This module is
the template and the reader. It touches no model and issues no query -- what it
produces is a list of rows and a list of complaints, and
``services.upload_plants`` is what turns those into stock.

That split is the point. Parsing a spreadsheet and deciding whether a strain
exists are different kinds of work with different failure modes, and keeping the
first free of the ORM means every coercion and every refusal below is testable
against a workbook built in memory.

**The cultivator is not a column, and the brief says it should be.**
``cultivator-stock-upload.md`` lists "Cultivator ID" among the required fields.
It is deliberately absent here: a column naming the cultivator is a column one
cultivator can fill in with another's identity, and the upload would then load
stock into somebody else's inventory. Who is uploading is a property of the
request and never of the file: the management command takes it as an argument,
and ``POST /api/stock/uploads`` takes it as a form field that ``plant.stock``
checks against the caller's own ``ProducerMembership`` rows before a single cell
is read. The brief's field list is describing a form a human fills in, not a
trust boundary.

**Dates must be dates, and a string is refused rather than guessed.**
``03/04/2026`` is the third of April in Johannesburg and the fourth of March in
Chicago, and a planting date read wrong by a month is a harvest estimate wrong by
a month that nobody notices until a member asks. So a date cell has to *be* a
date -- Excel's own type, which carries no ambiguity -- or an ISO ``YYYY-MM-DD``
string, which carries none either. Anything else is an error naming the fix.

**Money is refused rather than rounded.** A grow price of ``100.005`` is a typo
or a formula artefact, and quietly turning it into ``100.01`` puts a cent nobody
authorised onto a member's bill. Two decimal places or an error.
"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

#: The sheet plants are read from. Named rather than "the first sheet" so that a
#: cultivator who adds a working sheet of their own does not break the upload.
PLANTS_SHEET = 'Plants'

#: Where the template lists what the cultivator may choose from. Read by nothing
#: -- it exists so the person filling the form does not have to guess a strain
#: name, and so the dropdown on the strain column has somewhere to point.
REFERENCE_SHEET = 'Reference'

#: How many data rows an upload will look at. A cap rather than a stream, because
#: the error report is held in memory and a spreadsheet with a million blank-ish
#: rows -- which Excel produces easily -- would otherwise be a memory problem
#: rather than a validation one. Exceeding it is an error, never a silent
#: truncation.
MAX_ROWS = 2000

#: Column headings, in the order `cultivator-stock-upload.md` lists the fields.
#: The heading is the contract: the reader finds columns by name, so a cultivator
#: may reorder them and a later version may add one without invalidating every
#: template already downloaded.
COLUMNS = (
    ('cultivator_plant_id', 'Cultivator plant ID', True),
    ('strain', 'Strain', True),
    ('grow_price', 'Grow price (R)', True),
    ('planting_date', 'Planting date', True),
    ('estimated_bloom_date', 'Estimated bloom date', True),
    ('estimated_harvest_date', 'Estimated harvest date', True),
    ('minimum_yield_grams', 'Minimum yield (g)', True),
    ('finished_product_types', 'Finished product types', False),
    ('batch', 'Crop / batch number', False),
)

HEADINGS = {key: heading for key, heading, _ in COLUMNS}
REQUIRED = tuple(key for key, _, required in COLUMNS if required)

DATE_KEYS = ('planting_date', 'estimated_bloom_date', 'estimated_harvest_date')
DECIMAL_KEYS = ('grow_price', 'minimum_yield_grams')

#: Two decimal places, and no more. See the module docstring on why more is an
#: error rather than something to round.
MONEY = Decimal('0.01')


class RowError:
    """One thing wrong with one field, in the terms its consumer uses.

    ``row`` is the number Excel shows in its own margin, not a zero-based index
    into anything -- a cultivator fixing an upload is looking at the file, and an
    error report they have to translate is an error report they will misread.

    ``key`` is the internal field name and ``column`` is the spreadsheet heading
    derived from it. Both, because the same error has two audiences: an upload
    report names the column somebody is looking at, and a form -- the admin's, or
    Block 9's endpoint -- has to attach the message to a field. Carrying only the
    heading would leave the form matching on display text.
    """

    __slots__ = ('row', 'key', 'value', 'message')

    def __init__(self, row, key, value, message):
        self.row = row
        self.key = key
        self.value = value
        self.message = message

    @property
    def column(self):
        """The spreadsheet heading, for a report a cultivator reads."""
        return HEADINGS.get(self.key, self.key)

    def __str__(self):
        where = f'Row {self.row}'
        if self.column:
            where += f', {self.column}'
        if self.value not in (None, ''):
            where += f' ({self.value!r})'
        return f'{where}: {self.message}'

    def __repr__(self):
        return f'<RowError {self}>'

    def __eq__(self, other):
        return isinstance(other, RowError) and str(self) == str(other)


class SheetError(Exception):
    """The workbook itself is unusable, so no row was looked at.

    Distinct from a list of :class:`RowError` on purpose: a missing column is not
    a row a cultivator can fix, it is the wrong file. Telling them "row 2: strain
    is required" nine hundred times when the real problem is that the Strain
    heading was renamed is a report that hides its own cause.
    """


def build_template(strains=None):
    """A workbook a cultivator fills in. Returns an openpyxl ``Workbook``.

    ``strains`` is an iterable of ``(name, product types as text)`` pairs -- the
    cultivator's own listed offerings. Given them, the Strain column becomes a
    dropdown and the Reference sheet says what each strain will be delivered as,
    which removes the single most common upload error before it is made: a
    misspelt or unlisted strain name.

    The template is generated per cultivator rather than published once as a
    static file for exactly that reason. A generic template would need the
    cultivator to type strain names from memory.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = PLANTS_SHEET

    heading_font = Font(bold=True, color='FFFFFF')
    heading_fill = PatternFill('solid', start_color='2F5233')
    optional_fill = PatternFill('solid', start_color='6E8B6E')

    for index, (key, heading, required) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=heading)
        cell.font = heading_font
        cell.fill = heading_fill if required else optional_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)

        column = sheet.column_dimensions[get_column_letter(index)]
        column.width = max(len(heading) + 4, 18)

        if key in DATE_KEYS:
            # On the *column*, not on the cells. Writing a number format into
            # two thousand empty cells per date column would materialise six
            # thousand of them: the file gets fat, and -- the part that actually
            # broke -- the sheet's last row becomes 2001, so anything appended
            # afterwards lands below the rows an upload reads. A column format
            # applies to every cell in the column that has no style of its own,
            # which is all of the ones a cultivator has yet to type in.
            column.number_format = 'yyyy-mm-dd'

    sheet.freeze_panes = 'A2'

    reference = workbook.create_sheet(REFERENCE_SHEET)
    reference['A1'] = 'Strain'
    reference['B1'] = 'Will be delivered as'
    reference['A1'].font = Font(bold=True)
    reference['B1'].font = Font(bold=True)
    reference.column_dimensions['A'].width = 32
    reference.column_dimensions['B'].width = 48

    rows = list(strains or ())
    for offset, (name, product_types) in enumerate(rows, start=2):
        reference.cell(row=offset, column=1, value=name)
        reference.cell(row=offset, column=2, value=product_types)

    reference['D1'] = 'Notes'
    reference['D1'].font = Font(bold=True)
    for offset, note in enumerate(_template_notes(), start=2):
        reference.cell(row=offset, column=4, value=note)
    reference.column_dimensions['D'].width = 96

    if rows:
        strain_column = list(HEADINGS).index('strain') + 1
        letter = get_column_letter(strain_column)
        validation = DataValidation(
            type='list',
            formula1=f"='{REFERENCE_SHEET}'!$A$2:$A${len(rows) + 1}",
            allow_blank=False,
            showDropDown=False,
        )
        validation.error = (
            'Pick a strain you have a listing for. The Reference sheet has the '
            'list; ask an administrator to add a strain you cannot find.'
        )
        validation.errorTitle = 'Not one of your strains'
        sheet.add_data_validation(validation)
        validation.add(f'{letter}2:{letter}{MAX_ROWS + 1}')

    return workbook


def _template_notes():
    """What the Reference sheet tells the person filling the form.

    In the file rather than in an email, because the file is what gets forwarded
    to whoever actually does the typing.
    """
    return (
        'One row per plant. The dark green columns are required; the lighter '
        'ones are optional.',
        'Dates must be real dates, not text. Format the cell as a date, or type '
        'it as 2026-03-01 — a date like 03/04/2026 is ambiguous and will be '
        'refused rather than guessed at.',
        'Grow price and minimum yield are numbers, to at most two decimal '
        'places.',
        'Finished product types are optional. Leave the column blank and the '
        'plant offers whatever your strain listing offers. Fill it in and it '
        'must match that listing — this column confirms what you expect, it '
        'does not change it.',
        'Crop / batch number is optional and free text. Rows sharing one are '
        'grouped into the same batch.',
        'The platform allocates the serial, the leaf rating and the status. '
        'There are no columns for them.',
        'Nothing is loaded unless every row is valid. Fix what the report names '
        'and upload the same file again.',
    )


#: The export's columns, in the order a cultivator reads them. Distinct from
#: :data:`COLUMNS` on purpose, and almost its mirror image: the template omits
#: everything the platform generates because a cultivator must not supply it,
#: and the export *leads* with those, because they are the part a cultivator
#: cannot work out for themselves.
#:
#: Each entry is ``(key, heading, number format or None)``.
EXPORT_COLUMNS = (
    ('serial', 'Serial', None),
    ('cultivator_plant_id', 'Cultivator plant ID', None),
    ('strain', 'Strain', None),
    ('batch', 'Crop / batch', None),
    ('status', 'Status', None),
    ('leaf_rating', 'Leaf rating', '0.0'),
    ('grow_price', 'Grow price (R)', '#,##0.00'),
    ('minimum_yield_grams', 'Minimum yield (g)', '#,##0.00'),
    ('planting_date', 'Planted', 'yyyy-mm-dd'),
    ('estimated_bloom_date', 'Bloom (est.)', 'yyyy-mm-dd'),
    ('estimated_harvest_date', 'Harvest (est.)', 'yyyy-mm-dd'),
    ('harvested_on', 'Harvested', 'yyyy-mm-dd'),
    ('days_to_bloom', 'Days to bloom', '0'),
    ('days_to_harvest', 'Days to harvest', '0'),
    ('overdue', 'Overdue', None),
    ('finished_product_types', 'Delivered as', None),
    ('held_by', 'Held by', None),
)

#: The sheet stock is written to.
STOCK_SHEET = 'Stock'

#: Where the counts go. A dump of rows answers "what do I have"; a farm also
#: wants "how much, and how much of it is late" without writing a pivot table.
SUMMARY_SHEET = 'Summary'


def build_export(rows, summary=None, *, scope_label='', held_by=False):
    """A workbook of stock. Returns an openpyxl ``Workbook``.

    ``rows`` are mappings of **plain values** keyed by :data:`EXPORT_COLUMNS` --
    no model instances. That is what keeps this module's promise of touching no
    ORM, and it is ``services.stock_rows`` that flattens a queryset into them.
    The split is the same one the reader makes, for the same reason: the
    spreadsheet mechanics are testable without a database and the queries are
    testable without a spreadsheet.

    ``held_by`` includes the owner column. Off by default, and the default is the
    point: stock on hand is unsold by definition, so the column would be empty --
    and once it is not empty it carries a member's identity to a cultivator, which
    is a decision (see ``services.stock_rows``) rather than a formatting choice.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = STOCK_SHEET

    columns = [
        column for column in EXPORT_COLUMNS
        if held_by or column[0] != 'held_by'
    ]

    heading_font = Font(bold=True, color='FFFFFF')
    heading_fill = PatternFill('solid', start_color='2F5233')

    for index, (key, heading, number_format) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=heading)
        cell.font = heading_font
        cell.fill = heading_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)

        dimension = sheet.column_dimensions[get_column_letter(index)]
        dimension.width = max(len(heading) + 4, 14)
        if number_format:
            # On the column, never on two thousand cells. See `build_template`
            # for what that mistake cost.
            dimension.number_format = number_format

    for row in rows:
        sheet.append([row.get(key) for key, _, _ in columns])

    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = (
        f'A1:{get_column_letter(len(columns))}{max(sheet.max_row, 1)}'
    )

    _write_summary(workbook.create_sheet(SUMMARY_SHEET), summary, scope_label)
    return workbook


def _write_summary(sheet, summary, scope_label):
    sheet.column_dimensions['A'].width = 34
    sheet.column_dimensions['B'].width = 18

    def heading(row, text):
        cell = sheet.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True)

    row = 1
    heading(row, 'Stock on hand')
    row += 1
    if scope_label:
        sheet.cell(row=row, column=1, value='Scope')
        sheet.cell(row=row, column=2, value=scope_label)
        row += 2

    if not summary:
        return

    sheet.cell(row=row, column=1, value='Plants')
    sheet.cell(row=row, column=2, value=summary.get('plants', 0))
    row += 1
    sheet.cell(row=row, column=1, value='Total grow price (R)')
    total = sheet.cell(row=row, column=2, value=summary.get('grow_price_total', 0))
    total.number_format = '#,##0.00'
    row += 1
    sheet.cell(row=row, column=1, value='Minimum yield (g)')
    yield_cell = sheet.cell(row=row, column=2, value=summary.get('yield_total', 0))
    yield_cell.number_format = '#,##0.00'
    row += 1
    # "Late items" is one of the things the cultivator story asks the inventory
    # screen for by name, so it is on the summary rather than left to be
    # filtered for.
    sheet.cell(row=row, column=1, value='Overdue (past estimated harvest)')
    sheet.cell(row=row, column=2, value=summary.get('overdue', 0))
    row += 2

    for title, counts in (
        ('By status', summary.get('by_status') or ()),
        ('By strain', summary.get('by_strain') or ()),
    ):
        heading(row, title)
        row += 1
        for label, count in counts:
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=count)
            row += 1
        row += 1


def read_rows(source):
    """Read a filled-in template. Returns ``(rows, errors)``.

    ``rows`` are dicts keyed by the internal names in :data:`COLUMNS`, each
    carrying the spreadsheet row number as ``_row`` so that a later database-level
    complaint can still point at a line in the file. ``errors`` are
    :class:`RowError`.

    A row that fails coercion is left out of ``rows`` entirely rather than
    included half-parsed. ``services.upload_plants`` refuses to write anything
    when there are errors at all, so the only thing a partial row could do is
    produce a second, confusing complaint about a value that was never read.

    Raises :class:`SheetError` if the workbook is the wrong shape.
    """
    workbook = None
    try:
        try:
            workbook = load_workbook(source, read_only=True, data_only=True)
        except Exception as error:
            # openpyxl raises several unrelated types for "this is not a
            # spreadsheet" -- a zip error, a KeyError on a missing part, its own
            # InvalidFileException. A cultivator who uploaded a .csv or a PDF
            # needs one sentence, not the distinction between them.
            raise SheetError(
                'That file could not be opened as an Excel workbook. Save it as '
                '.xlsx and upload it again.'
            ) from error

        if PLANTS_SHEET not in workbook.sheetnames:
            raise SheetError(
                f'The workbook has no {PLANTS_SHEET!r} sheet. Use the template '
                'from `manage.py plant_template` rather than a new file.'
            )

        sheet = workbook[PLANTS_SHEET]
        return _read_sheet(sheet)
    finally:
        if workbook is not None:
            workbook.close()


def _read_sheet(sheet):
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise SheetError('The Plants sheet is empty.') from None

    index_of = _column_index(header)

    rows = []
    errors = []
    seen_plant_ids = {}

    for offset, values in enumerate(rows_iter, start=2):
        if offset - 1 > MAX_ROWS:
            raise SheetError(
                f'That workbook has more than {MAX_ROWS} rows of plants. Split '
                'it and upload the parts separately -- nothing was loaded.'
            )

        if _is_blank(values):
            # Excel hands back trailing empty rows for any sheet somebody has
            # scrolled through. Skipped rather than reported, or every upload
            # would carry hundreds of complaints about rows nobody typed.
            continue

        raw = {key: _cell(values, position) for key, position in index_of.items()}
        row, row_errors = read_row(raw, row_number=offset)

        # Within the file as well as against the database. Two rows claiming one
        # plant ID is a mistake the unique constraint would catch on the second
        # insert, halfway through a batch, with a message naming an index.
        if row is not None:
            plant_id = row['cultivator_plant_id']
            key = (row['strain'].casefold(), plant_id.casefold())
            if key in seen_plant_ids:
                row_errors.append(RowError(
                    offset,
                    'cultivator_plant_id',
                    plant_id,
                    f'Already used for this strain on row {seen_plant_ids[key]}.',
                ))
                row = None
            else:
                seen_plant_ids[key] = offset

        if row is None:
            errors.extend(row_errors)
        else:
            errors.extend(row_errors)
            rows.append(row)

    if not rows and not errors:
        raise SheetError(
            'That workbook has headings and no plants. Fill in a row per plant '
            'and upload it again.'
        )

    return rows, errors


def _column_index(header):
    """Map internal names to column positions, by heading.

    By name and not by position, so a cultivator may reorder columns or leave a
    stray one in place. A missing *required* heading is a `SheetError`: it is the
    wrong file, not a bad row.
    """
    found = {}
    for position, value in enumerate(header or ()):
        if value is None:
            continue
        label = str(value).strip().casefold()
        for key, heading, _ in COLUMNS:
            if label == heading.casefold():
                found[key] = position

    missing = [HEADINGS[key] for key in REQUIRED if key not in found]
    if missing:
        raise SheetError(
            'The workbook is missing these columns: '
            + ', '.join(missing)
            + '. Use the template from `manage.py plant_template`.'
        )
    return found


def _is_blank(values):
    return all(value is None or str(value).strip() == '' for value in values or ())


def read_row(raw, row_number=1):
    """Coerce one plant's raw values. Returns ``(row or None, errors)``.

    Public, and the reason this module has a single-row entry point at all: an
    upload and an individual capture are the same nine fields with two different
    sources, and ``cultivator-stock-upload.md`` describes them as one list of
    requirements. Two implementations of "is this a date" would eventually
    disagree, and the one that disagreed would be whichever was used less.

    ``raw`` is a mapping keyed by the internal names in :data:`COLUMNS`, holding
    whatever the source produced -- an Excel cell, a form field, a JSON value.
    Missing keys are treated as blank, so a caller need not supply the optional
    ones.

    ``row_number`` is what the errors are labelled with. It defaults to 1 for a
    single capture, where there is no row to point at and the number is noise;
    the sheet reader passes the line the cultivator is looking at.
    """
    errors = []

    for key in REQUIRED:
        if raw.get(key) in (None, ''):
            errors.append(RowError(
                row_number, key, None, 'This is required.'
            ))

    row = {'_row': row_number}

    for key in ('cultivator_plant_id', 'strain', 'batch'):
        row[key] = str(raw.get(key) or '').strip()

    for key in DECIMAL_KEYS:
        value = raw.get(key)
        if value in (None, ''):
            continue
        amount, error = _to_decimal(value, row_number, key)
        if error:
            errors.append(error)
        else:
            row[key] = amount

    for key in DATE_KEYS:
        value = raw.get(key)
        if value in (None, ''):
            continue
        when, error = _to_date(value, row_number, key)
        if error:
            errors.append(error)
        else:
            row[key] = when

    row['finished_product_types'] = _to_list(raw.get('finished_product_types'))

    if errors:
        return None, errors

    # Cheap ordering checks here rather than in the service. They need no query,
    # and a cultivator who has transposed two dates would otherwise be told about
    # it by a check constraint naming an index.
    if row['estimated_bloom_date'] < row['planting_date']:
        errors.append(RowError(
            row_number,
            'estimated_bloom_date',
            row['estimated_bloom_date'].isoformat(),
            'A plant cannot flower before it was planted.',
        ))
    if row['estimated_harvest_date'] < row['planting_date']:
        errors.append(RowError(
            row_number,
            'estimated_harvest_date',
            row['estimated_harvest_date'].isoformat(),
            'A plant cannot be harvested before it was planted.',
        ))
    if row['estimated_harvest_date'] < row['estimated_bloom_date']:
        errors.append(RowError(
            row_number,
            'estimated_harvest_date',
            row['estimated_harvest_date'].isoformat(),
            'Harvest is expected before bloom. Check the two dates.',
        ))

    if errors:
        return None, errors
    return row, []


def _cell(values, position):
    if position >= len(values or ()):
        return None
    value = values[position]
    if isinstance(value, str):
        value = value.strip()
    return value or None if isinstance(value, str) else value


def _to_decimal(value, row_number, key):
    if isinstance(value, bool):
        # `True` is an int in Python and would quietly become R1.
        return None, RowError(row_number, key, value, 'This must be a number.')
    try:
        amount = Decimal(str(value).replace(' ', '').replace(',', ''))
    except (InvalidOperation, ValueError, TypeError):
        return None, RowError(
            row_number, key, value, 'This must be a number.'
        )

    if amount != amount.quantize(MONEY):
        return None, RowError(
            row_number, key, value,
            'At most two decimal places. This has more, which is usually a '
            'formula rather than a price.',
        )
    if amount <= 0:
        return None, RowError(
            row_number, key, value, 'This must be more than zero.'
        )
    return amount.quantize(MONEY), None


def _to_date(value, row_number, key):
    if isinstance(value, datetime):
        return value.date(), None
    if isinstance(value, date):
        return value, None
    try:
        return date.fromisoformat(str(value).strip()), None
    except (ValueError, TypeError):
        return None, RowError(
            row_number, key, value,
            'This must be a date. Format the cell as a date, or type it as '
            '2026-03-01 — a date like 03/04/2026 means different things in '
            'different places and will not be guessed at.',
        )


def _to_list(value):
    if value in (None, ''):
        return []
    return [part.strip() for part in str(value).split(',') if part.strip()]
